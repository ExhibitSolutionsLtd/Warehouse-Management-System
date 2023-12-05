import openpyxl
from .models import Product, Location, Customer, Supplier


def import_from_excel(file, user):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active


    for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming first row is header
        zone = row[3]
        item_row = row[4]
        bay = row[5]
        tier = row[6]
        location, created = Location.objects.get_or_create(zone = zone, row = item_row, bay = bay, tier = tier)
        

        product = Product(sku=row[0], item_name=row[1], category=row[2], quantity=row[7], description=row[8], location=location, user=user)
        product.save()

def import_customer(file):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming first row is header
        
        customer = Customer(cust_f_name=row[0], cust_l_name=row[1], email=row[2], mobile_no=row[3], address=row[4], notes=row[5])
        customer.save()


def import_supplier(file):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming first row is header
        
        supplier = Supplier(sup_f_name=row[0], sup_l_name=row[1], email=row[2], mobile_no=row[3], address=row[4], notes=row[5])
        supplier.save()